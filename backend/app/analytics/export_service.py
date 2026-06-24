"""
Analytics export service — read-only, tenant-scoped.
Generates CSV or XLSX bytes for browser download.
"""
from __future__ import annotations

import csv
import io
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from sqlalchemy import and_, func, literal, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.analytics.financial_reports import FinancialReportsService
from app.analytics.inventory_reports import InventoryReportsService
from app.analytics.sales_reports import SalesReportsService
from app.cashiers.models import CashierSession
from app.core.constants import OrderStatus, PaymentStatus
from app.customers.models import Customer
from app.models.branch import Branch
from app.models.inventory import BranchInventory, StockMovement
from app.models.product import Category, Product, ProductVariant
from app.models.supplier import Supplier
from app.models.user import User
from app.payments.models import Payment, Refund, RefundItem
from app.procurement.models import (
    GoodsReceipt,
    GoodsReceiptItem,
    PurchaseOrder,
    PurchaseOrderItem,
    SupplierPayable,
)
from app.sales.models import Order, OrderItem


def _utc_start(d: date) -> datetime:
    return datetime(d.year, d.month, d.day, tzinfo=timezone.utc)


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def _fmt_dec(v: Decimal | None) -> str:
    if v is None:
        return ""
    return str(v.quantize(Decimal("0.00")))


# Columns whose values are numeric and should be summed in the TOTAL row.
_SUM_COLS: frozenset[str] = frozenset({
    # Sales / refunds
    "Subtotal", "Discount", "Tax", "Total", "Refunded Amount", "Net Amount",
    "Unit Price", "Line Refund Amount", "Total Refund Amount",
    "Unit Cost", "Line Subtotal", "Line Total", "Order Total",
    # Inventory stocks
    "On Hand", "Reserved", "Available", "Reorder Point", "Reorder Qty",
    "Stock Value",
    # Cashier performance
    "Gross Sales", "Refunds", "Net Sales", "Avg Order Value", "Orders",
    # Top products & category
    "Units Sold", "Revenue", "Avg Unit Price", "Profit Estimate",
    # Payment methods
    "Total Amount", "Transactions",
    # Trend
    "Net Revenue",
    # Profit report
    "COGS", "Gross Profit",
    # Low stock
    "Shortage",
    # Stock movements
    "Qty Change", "Movement Value",
    # Customers
    "Balance", "Credit Limit", "Total Spent",
    # Procurement
    "Invoice Amount", "Paid Amount", "Remaining Amount",
    "Ordered Qty", "Received Qty",
})


def _totals_row(headers: list[str], rows: list[dict]) -> dict:
    """Build a TOTAL row that sums every numeric column; non-numeric cols are blank."""
    totals: dict = {h: "" for h in headers}
    totals[headers[0]] = "TOTAL"
    for h in headers:
        if h in _SUM_COLS:
            total = Decimal("0")
            for row in rows:
                val = row.get(h, "")
                if val:
                    try:
                        total += Decimal(val)
                    except Exception:
                        pass
            totals[h] = str(total.quantize(Decimal("0.00")))
    return totals


def _write_section(buf: io.StringIO, title: str, headers: list[str], rows: list[dict]) -> None:
    """Write a titled section (with a TOTAL row) into an already-open StringIO buffer."""
    buf.write(f"{title}\n")
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    if rows:
        writer.writerow(_totals_row(headers, rows))
    # blank separator row between sections
    buf.write("\n")


def _build_csv(*sections: tuple[str, list[str], list[dict]]) -> bytes:
    """Combine multiple (title, headers, rows) sections into one CSV file."""
    buf = io.StringIO()
    for title, headers, rows in sections:
        _write_section(buf, title, headers, rows)
    # UTF-8 BOM → Excel / Google Sheets open without encoding dialog
    return "﻿".encode("utf-8") + buf.getvalue().encode("utf-8")


def _coerce_xlsx(val: str) -> float | str:
    """Convert numeric strings to float so Excel treats them as numbers."""
    if not val or val == "TOTAL":
        return val
    try:
        return float(val)
    except (ValueError, TypeError):
        return val


def _build_xlsx(*sections: tuple[str, list[str], list[dict]]) -> bytes:
    """Combine multiple (title, headers, rows) sections into an XLSX workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    HDR_FILL = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
    HDR_FONT = Font(bold=True, color="F59E0B", size=10)
    HDR_BORDER = Border(bottom=Side(border_style="medium", color="F59E0B"))
    TOTAL_FILL = PatternFill(start_color="3F2800", end_color="3F2800", fill_type="solid")
    TOTAL_FONT = Font(bold=True, color="FCD34D", size=10)
    DATA_FONT = Font(size=10)

    for title, headers, rows in sections:
        ws = wb.create_sheet(title=title[:31])

        # Header row
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = HDR_BORDER
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Data rows
        for ri, row in enumerate(rows, 2):
            for ci, header in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=_coerce_xlsx(row.get(header, "")))
                cell.font = DATA_FONT

        # TOTAL row
        if rows:
            total_row = _totals_row(headers, rows)
            tri = len(rows) + 2
            for ci, header in enumerate(headers, 1):
                cell = ws.cell(row=tri, column=ci, value=_coerce_xlsx(total_row.get(header, "")))
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=8,
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    # shared helpers

    def _order_filters(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None,
        end_date: date | None,
        branch_id: uuid.UUID | None,
    ) -> list:
        f: list = [
            Order.tenant_id == tenant_id,
            Order.order_status.in_([
                OrderStatus.COMPLETED.value,
                OrderStatus.PARTIALLY_REFUNDED.value,
                OrderStatus.REFUNDED.value,
            ]),
        ]
        if start_date:
            f.append(Order.created_at >= _utc_start(start_date))
        if end_date:
            f.append(Order.created_at < _utc_start(end_date) + timedelta(days=1))
        if branch_id:
            f.append(Order.branch_id == branch_id)
        return f

    def _pay_methods_subq(self, tenant_id: uuid.UUID):
        """Subquery: order_id → comma-separated list of distinct payment methods."""
        return (
            select(
                Payment.order_id.label("order_id"),
                func.string_agg(Payment.payment_method, literal(", ")).label("methods"),
            )
            .where(
                and_(
                    Payment.tenant_id == tenant_id,
                    Payment.payment_status == PaymentStatus.PAID.value,
                )
            )
            .group_by(Payment.order_id)
            .subquery("pay_methods")
        )

    # export 1: sales + refunds combined

    async def export_sales_and_refunds(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        """
        Single CSV with two sections:
          SECTION 1 — SALES: one row per order
          SECTION 2 — REFUNDS: one row per refund line item
        """
        sales_rows = await self._fetch_sales(tenant_id, start_date, end_date, branch_id)
        refund_rows = await self._fetch_refunds(tenant_id, start_date, end_date, branch_id)

        sales_headers = [
            "Order Number", "Date", "Branch", "Cashier", "Customer",
            "Subtotal", "Discount", "Tax", "Total",
            "Payment Methods", "Status",
            "Refunded Amount", "Net Amount",
            "Notes", "Completed At",
        ]
        refund_headers = [
            "Refund Number", "Refund Date",
            "Original Order", "Order Date",
            "Branch", "Customer",
            "Product", "Variant",
            "Qty", "Unit Price", "Line Refund Amount",
            "Total Refund Amount",
            "Reason", "Type", "Processed By", "Notes",
        ]

        sections = (
            ("SALES", sales_headers, sales_rows),
            ("REFUNDS", refund_headers, refund_rows),
        )
        return _build_xlsx(*sections) if fmt == "xlsx" else _build_csv(*sections)

    async def _fetch_sales(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None,
        end_date: date | None,
        branch_id: uuid.UUID | None,
    ) -> list[dict]:
        CashierUser = aliased(User)
        pay_subq = self._pay_methods_subq(tenant_id)

        stmt = (
            select(
                Order.order_number,
                Order.created_at,
                Branch.name.label("branch_name"),
                func.concat(CashierUser.first_name, literal(" "), CashierUser.last_name).label("cashier_name"),
                Customer.name.label("customer_name"),
                Order.subtotal,
                Order.discount_amount,
                Order.tax_amount,
                Order.total_amount,
                pay_subq.c.methods.label("payment_methods"),
                Order.order_status,
                Order.refunded_amount,
                (Order.total_amount - Order.refunded_amount).label("net_amount"),
                Order.notes,
                Order.completed_at,
            )
            .join(Branch, Branch.id == Order.branch_id)
            .join(CashierSession, CashierSession.id == Order.cashier_session_id)
            .join(CashierUser, CashierUser.id == CashierSession.cashier_user_id)
            .outerjoin(Customer, Customer.id == Order.customer_id)
            .outerjoin(pay_subq, pay_subq.c.order_id == Order.id)
            .where(and_(*self._order_filters(tenant_id, start_date, end_date, branch_id)))
            .order_by(Order.created_at.desc())
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()
        return [
            {
                "Order Number":    r.order_number,
                "Date":            _fmt_dt(r.created_at),
                "Branch":          r.branch_name or "",
                "Cashier":         r.cashier_name or "",
                "Customer":        r.customer_name or "",
                "Subtotal":        _fmt_dec(r.subtotal),
                "Discount":        _fmt_dec(r.discount_amount),
                "Tax":             _fmt_dec(r.tax_amount),
                "Total":           _fmt_dec(r.total_amount),
                "Payment Methods": r.payment_methods or "",
                "Status":          r.order_status,
                "Refunded Amount": _fmt_dec(r.refunded_amount),
                "Net Amount":      _fmt_dec(r.net_amount),
                "Notes":           r.notes or "",
                "Completed At":    _fmt_dt(r.completed_at),
            }
            for r in rows_raw
        ]

    async def _fetch_refunds(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None,
        end_date: date | None,
        branch_id: uuid.UUID | None,
    ) -> list[dict]:
        ProcessedByUser = aliased(User)

        refund_filters: list = [Refund.tenant_id == tenant_id]
        if start_date:
            refund_filters.append(Refund.processed_at >= _utc_start(start_date))
        if end_date:
            refund_filters.append(Refund.processed_at < _utc_start(end_date) + timedelta(days=1))
        if branch_id:
            refund_filters.append(Order.branch_id == branch_id)

        stmt = (
            select(
                Refund.refund_number,
                Refund.processed_at,
                Order.order_number,
                Order.created_at.label("order_date"),
                Branch.name.label("branch_name"),
                Customer.name.label("customer_name"),
                OrderItem.product_name,
                OrderItem.variant_name,
                RefundItem.quantity,
                OrderItem.unit_price,
                RefundItem.amount.label("line_refund_amount"),
                Refund.amount.label("total_refund_amount"),
                Refund.reason,
                Refund.refund_type,
                func.concat(ProcessedByUser.first_name, literal(" "), ProcessedByUser.last_name).label("processed_by"),
                Refund.notes,
            )
            .join(Order, Order.id == Refund.order_id)
            .join(Branch, Branch.id == Order.branch_id)
            .outerjoin(Customer, Customer.id == Order.customer_id)
            .join(ProcessedByUser, ProcessedByUser.id == Refund.processed_by)
            .outerjoin(RefundItem, RefundItem.refund_id == Refund.id)
            .outerjoin(OrderItem, OrderItem.id == RefundItem.order_item_id)
            .where(and_(*refund_filters))
            .order_by(Refund.processed_at.desc(), Refund.refund_number)
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()
        return [
            {
                "Refund Number":       r.refund_number,
                "Refund Date":         _fmt_dt(r.processed_at),
                "Original Order":      r.order_number,
                "Order Date":          _fmt_dt(r.order_date),
                "Branch":              r.branch_name or "",
                "Customer":            r.customer_name or "",
                "Product":             r.product_name or "",
                "Variant":             r.variant_name or "",
                "Qty":                 str(r.quantity) if r.quantity is not None else "",
                "Unit Price":          _fmt_dec(r.unit_price) if r.unit_price is not None else "",
                "Line Refund Amount":  _fmt_dec(r.line_refund_amount) if r.line_refund_amount is not None else "",
                "Total Refund Amount": _fmt_dec(r.total_refund_amount),
                "Reason":              r.reason,
                "Type":                r.refund_type,
                "Processed By":        r.processed_by or "",
                "Notes":               r.notes or "",
            }
            for r in rows_raw
        ]

    # export 2: order line items

    async def export_order_items(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        """
        One row per order line item — ideal for COGS / margin analysis.
        """
        CashierUser = aliased(User)
        pay_subq = self._pay_methods_subq(tenant_id)

        stmt = (
            select(
                Order.order_number,
                Order.created_at,
                Branch.name.label("branch_name"),
                func.concat(CashierUser.first_name, literal(" "), CashierUser.last_name).label("cashier_name"),
                Customer.name.label("customer_name"),
                OrderItem.product_name,
                OrderItem.variant_name,
                OrderItem.sku,
                OrderItem.quantity,
                OrderItem.unit_price,
                OrderItem.unit_cost_snapshot,
                OrderItem.discount_amount,
                OrderItem.tax_rate,
                OrderItem.subtotal.label("line_subtotal"),
                OrderItem.total.label("line_total"),
                Order.total_amount.label("order_total"),
                pay_subq.c.methods.label("payment_methods"),
                Order.order_status,
            )
            .join(OrderItem, OrderItem.order_id == Order.id)
            .join(Branch, Branch.id == Order.branch_id)
            .join(CashierSession, CashierSession.id == Order.cashier_session_id)
            .join(CashierUser, CashierUser.id == CashierSession.cashier_user_id)
            .outerjoin(Customer, Customer.id == Order.customer_id)
            .outerjoin(pay_subq, pay_subq.c.order_id == Order.id)
            .where(and_(*self._order_filters(tenant_id, start_date, end_date, branch_id)))
            .order_by(Order.created_at.desc(), OrderItem.id)
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()

        headers = [
            "Order Number", "Order Date", "Branch", "Cashier", "Customer",
            "Product", "Variant", "SKU",
            "Qty", "Unit Price", "Unit Cost",
            "Discount", "Tax Rate",
            "Line Subtotal", "Line Total", "Order Total",
            "Payment Methods", "Order Status",
        ]

        rows = [
            {
                "Order Number":    r.order_number,
                "Order Date":      _fmt_dt(r.created_at),
                "Branch":          r.branch_name or "",
                "Cashier":         r.cashier_name or "",
                "Customer":        r.customer_name or "",
                "Product":         r.product_name,
                "Variant":         r.variant_name or "",
                "SKU":             r.sku or "",
                "Qty":             str(r.quantity),
                "Unit Price":      _fmt_dec(r.unit_price),
                "Unit Cost":       _fmt_dec(r.unit_cost_snapshot) if r.unit_cost_snapshot is not None else "",
                "Discount":        _fmt_dec(r.discount_amount),
                "Tax Rate":        str(r.tax_rate),
                "Line Subtotal":   _fmt_dec(r.line_subtotal),
                "Line Total":      _fmt_dec(r.line_total),
                "Order Total":     _fmt_dec(r.order_total),
                "Payment Methods": r.payment_methods or "",
                "Order Status":    r.order_status,
            }
            for r in rows_raw
        ]

        return _build_xlsx(("ORDERS", headers, rows)) if fmt == "xlsx" else _build_csv(("ORDERS", headers, rows))

    # export 3: current inventory stocks

    async def export_inventory_stocks(
        self,
        tenant_id: uuid.UUID,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        """
        Snapshot of current stock levels — one row per product/variant per branch.
        """
        stmt = (
            select(
                Branch.name.label("branch_name"),
                Category.name.label("category_name"),
                Product.name.label("product_name"),
                ProductVariant.name.label("variant_name"),
                func.coalesce(ProductVariant.sku, Product.sku).label("sku"),
                BranchInventory.quantity_on_hand,
                BranchInventory.quantity_reserved,
                (BranchInventory.quantity_on_hand - BranchInventory.quantity_reserved).label("quantity_available"),
                BranchInventory.reorder_point,
                BranchInventory.reorder_quantity,
                func.coalesce(ProductVariant.cost_price, Product.cost_price).label("unit_cost"),
                (
                    BranchInventory.quantity_on_hand
                    * func.coalesce(ProductVariant.cost_price, Product.cost_price, literal(Decimal("0")))
                ).label("stock_value"),
                BranchInventory.last_movement_at,
            )
            .join(Branch, Branch.id == BranchInventory.branch_id)
            .join(Product, Product.id == BranchInventory.product_id)
            .outerjoin(ProductVariant, ProductVariant.id == BranchInventory.variant_id)
            .outerjoin(Category, Category.id == Product.category_id)
            .where(
                and_(
                    BranchInventory.tenant_id == tenant_id,
                    Product.is_deleted.is_(False),
                    *([BranchInventory.branch_id == branch_id] if branch_id else []),
                )
            )
            .order_by(Branch.name, Product.name, ProductVariant.name)
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()

        headers = [
            "Branch", "Category", "Product", "Variant", "SKU",
            "On Hand", "Reserved", "Available",
            "Reorder Point", "Reorder Qty",
            "Unit Cost", "Stock Value",
            "Last Movement",
        ]

        rows = [
            {
                "Branch":        r.branch_name or "",
                "Category":      r.category_name or "",
                "Product":       r.product_name,
                "Variant":       r.variant_name or "",
                "SKU":           r.sku or "",
                "On Hand":       _fmt_dec(r.quantity_on_hand),
                "Reserved":      _fmt_dec(r.quantity_reserved),
                "Available":     _fmt_dec(r.quantity_available),
                "Reorder Point": _fmt_dec(r.reorder_point) if r.reorder_point is not None else "",
                "Reorder Qty":   _fmt_dec(r.reorder_quantity) if r.reorder_quantity is not None else "",
                "Unit Cost":     _fmt_dec(r.unit_cost) if r.unit_cost is not None else "",
                "Stock Value":   _fmt_dec(r.stock_value) if r.stock_value is not None else "",
                "Last Movement": _fmt_dt(r.last_movement_at),
            }
            for r in rows_raw
        ]

        return _build_xlsx(("INVENTORY STOCKS", headers, rows)) if fmt == "xlsx" else _build_csv(("INVENTORY STOCKS", headers, rows))

    # export 4: top products by revenue

    async def export_top_products(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = SalesReportsService(self.session)
        items = await svc.get_top_products(
            tenant_id, start_date, end_date, branch_id, limit=500
        )
        headers = [
            "Rank", "Product", "SKU",
            "Units Sold", "Revenue", "Avg Unit Price", "Profit Estimate", "Margin %",
        ]
        rows = []
        for rank, item in enumerate(items, 1):
            avg_price = (
                (item.revenue / item.quantity_sold).quantize(Decimal("0.00"))
                if item.quantity_sold else Decimal("0")
            )
            margin = (
                (item.profit_estimate / item.revenue * 100).quantize(Decimal("0.00"))
                if item.revenue else Decimal("0")
            )
            rows.append({
                "Rank": str(rank),
                "Product": item.product_name,
                "SKU": item.sku or "",
                "Units Sold": _fmt_dec(item.quantity_sold),
                "Revenue": _fmt_dec(item.revenue),
                "Avg Unit Price": _fmt_dec(avg_price),
                "Profit Estimate": _fmt_dec(item.profit_estimate),
                "Margin %": _fmt_dec(margin),
            })
        return (
            _build_xlsx(("TOP PRODUCTS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("TOP PRODUCTS", headers, rows))
        )

    # export 5: sales by cashier / staff performance

    async def export_sales_by_cashier(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = SalesReportsService(self.session)
        items = await svc.get_by_cashier(tenant_id, start_date, end_date, branch_id)
        headers = ["Cashier", "Orders", "Gross Sales", "Refunds", "Net Sales", "Avg Order Value"]
        rows = [
            {
                "Cashier": item.cashier_name,
                "Orders": str(item.orders),
                "Gross Sales": _fmt_dec(item.sales),
                "Refunds": _fmt_dec(item.refunds),
                "Net Sales": _fmt_dec(item.sales - item.refunds),
                "Avg Order Value": _fmt_dec(item.average_ticket),
            }
            for item in items
        ]
        return (
            _build_xlsx(("SALES BY CASHIER", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("SALES BY CASHIER", headers, rows))
        )

    # export 6: sales by category

    async def export_sales_by_category(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = SalesReportsService(self.session)
        items = await svc.get_by_category(tenant_id, start_date, end_date, branch_id)
        total_revenue = sum(item.sales for item in items) or Decimal("1")
        headers = ["Category", "Units Sold", "Revenue", "Share %", "Profit Estimate"]
        rows = [
            {
                "Category": item.category_name,
                "Units Sold": _fmt_dec(item.quantity_sold),
                "Revenue": _fmt_dec(item.sales),
                "Share %": _fmt_dec(
                    (item.sales / total_revenue * 100).quantize(Decimal("0.00"))
                ),
                "Profit Estimate": _fmt_dec(item.profit),
            }
            for item in items
        ]
        return (
            _build_xlsx(("SALES BY CATEGORY", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("SALES BY CATEGORY", headers, rows))
        )

    # export 7: payment methods breakdown

    async def export_payment_methods(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = SalesReportsService(self.session)
        items = await svc.get_payment_methods(tenant_id, start_date, end_date, branch_id)
        headers = ["Payment Method", "Transactions", "Total Amount", "Share %"]
        rows = [
            {
                "Payment Method": item.payment_method,
                "Transactions": str(item.transaction_count),
                "Total Amount": _fmt_dec(item.amount),
                "Share %": _fmt_dec(item.percentage),
            }
            for item in items
        ]
        return (
            _build_xlsx(("PAYMENT METHODS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("PAYMENT METHODS", headers, rows))
        )

    # export 8: sales trend

    async def export_sales_trend(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        granularity: str = "daily",
        fmt: str = "csv",
    ) -> bytes:
        svc = SalesReportsService(self.session)
        trend = await svc.get_trend(tenant_id, granularity, start_date, end_date, branch_id)
        headers = ["Period", "Orders", "Gross Sales", "Net Revenue"]
        rows = [
            {
                "Period": item.period,
                "Orders": str(item.orders),
                "Gross Sales": _fmt_dec(item.sales),
                "Net Revenue": _fmt_dec(item.revenue),
            }
            for item in trend.items
        ]
        title = f"SALES TREND ({granularity.upper()})"
        return (
            _build_xlsx((title, headers, rows))
            if fmt == "xlsx"
            else _build_csv((title, headers, rows))
        )

    # export 9: profit report (by product, category, branch — all sheets in one file)

    async def export_profit_report(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = FinancialReportsService(self.session)

        def _profit_rows(items, label_col: str) -> tuple[list[str], list[dict]]:
            headers = [label_col, "Revenue", "COGS", "Gross Profit", "Margin %"]
            rows = [
                {
                    label_col: item.dimension_name,
                    "Revenue": _fmt_dec(item.revenue),
                    "COGS": _fmt_dec(item.cogs),
                    "Gross Profit": _fmt_dec(item.profit),
                    "Margin %": _fmt_dec(item.margin_pct),
                }
                for item in items
            ]
            return headers, rows

        by_product = await svc.get_profit_report(tenant_id, "product", start_date, end_date, branch_id)
        by_category = await svc.get_profit_report(tenant_id, "category", start_date, end_date, branch_id)
        by_branch = await svc.get_profit_report(tenant_id, "branch", start_date, end_date)

        h_p, r_p = _profit_rows(by_product.items, "Product")
        h_c, r_c = _profit_rows(by_category.items, "Category")
        h_b, r_b = _profit_rows(by_branch.items, "Branch")

        sections = (
            ("BY PRODUCT", h_p, r_p),
            ("BY CATEGORY", h_c, r_c),
            ("BY BRANCH", h_b, r_b),
        )
        return _build_xlsx(*sections) if fmt == "xlsx" else _build_csv(*sections)

    # export 10: low stock items

    async def export_low_stock(
        self,
        tenant_id: uuid.UUID,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = InventoryReportsService(self.session)
        items = await svc.get_low_stock(tenant_id, branch_id)
        headers = ["Branch", "Product", "SKU", "On Hand", "Reorder Point", "Shortage"]
        rows = [
            {
                "Branch": item.branch_name,
                "Product": item.product_name,
                "SKU": item.sku or "",
                "On Hand": _fmt_dec(item.quantity_on_hand),
                "Reorder Point": _fmt_dec(item.reorder_point),
                "Shortage": _fmt_dec(item.reorder_point - item.quantity_on_hand),
            }
            for item in items
        ]
        return (
            _build_xlsx(("LOW STOCK ITEMS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("LOW STOCK ITEMS", headers, rows))
        )

    # export 11: fast-moving products

    async def export_fast_moving(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = InventoryReportsService(self.session)
        items = await svc.get_fast_moving(tenant_id, start_date, end_date, branch_id, limit=500)
        headers = ["Rank", "Product", "SKU", "Units Sold", "Order Count"]
        rows = [
            {
                "Rank": str(item.rank),
                "Product": item.product_name,
                "SKU": item.sku or "",
                "Units Sold": _fmt_dec(item.quantity_sold),
                "Order Count": str(item.order_count),
            }
            for item in items
        ]
        return (
            _build_xlsx(("FAST MOVING PRODUCTS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("FAST MOVING PRODUCTS", headers, rows))
        )

    # export 12: dead stock

    async def export_dead_stock(
        self,
        tenant_id: uuid.UUID,
        days: int = 90,
        branch_id: uuid.UUID | None = None,
        fmt: str = "csv",
    ) -> bytes:
        svc = InventoryReportsService(self.session)
        items = await svc.get_dead_stock(tenant_id, days, branch_id)
        headers = ["Product", "SKU", "On Hand", "Last Sold", "Days Without Sale"]
        rows = [
            {
                "Product": item.product_name,
                "SKU": item.sku or "",
                "On Hand": _fmt_dec(item.quantity_on_hand),
                "Last Sold": (
                    item.last_sold_at.strftime("%Y-%m-%d")
                    if item.last_sold_at
                    else "Never"
                ),
                "Days Without Sale": str(item.days_without_sale),
            }
            for item in items
        ]
        title = f"DEAD STOCK (>{days} DAYS)"
        return (
            _build_xlsx((title, headers, rows))
            if fmt == "xlsx"
            else _build_csv((title, headers, rows))
        )

    # export 13: individual stock movements

    async def export_stock_movements(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        branch_id: uuid.UUID | None = None,
        movement_type: str | None = None,
        fmt: str = "csv",
    ) -> bytes:
        filters: list = [StockMovement.tenant_id == tenant_id]
        if start_date:
            filters.append(StockMovement.created_at >= _utc_start(start_date))
        if end_date:
            filters.append(
                StockMovement.created_at < _utc_start(end_date) + timedelta(days=1)
            )
        if branch_id:
            filters.append(StockMovement.branch_id == branch_id)
        if movement_type:
            filters.append(StockMovement.movement_type == movement_type)

        stmt = (
            select(
                StockMovement.created_at.label("date"),
                Branch.name.label("branch_name"),
                Product.name.label("product_name"),
                ProductVariant.name.label("variant_name"),
                func.coalesce(ProductVariant.sku, Product.sku).label("sku"),
                StockMovement.movement_type,
                StockMovement.quantity.label("qty_change"),
                StockMovement.unit_cost,
                (
                    StockMovement.quantity
                    * func.coalesce(StockMovement.unit_cost, literal(Decimal("0")))
                ).label("movement_value"),
                StockMovement.reference_type,
                StockMovement.reason,
                StockMovement.notes,
            )
            .join(Branch, Branch.id == StockMovement.branch_id)
            .join(Product, Product.id == StockMovement.product_id)
            .outerjoin(ProductVariant, ProductVariant.id == StockMovement.variant_id)
            .where(and_(*filters))
            .order_by(StockMovement.created_at.desc())
            .limit(10000)
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()
        headers = [
            "Date", "Branch", "Product", "Variant", "SKU",
            "Movement Type", "Qty Change", "Unit Cost", "Movement Value",
            "Reference Type", "Reason", "Notes",
        ]
        rows = [
            {
                "Date": _fmt_dt(r.date),
                "Branch": r.branch_name or "",
                "Product": r.product_name,
                "Variant": r.variant_name or "",
                "SKU": r.sku or "",
                "Movement Type": r.movement_type,
                "Qty Change": _fmt_dec(r.qty_change),
                "Unit Cost": _fmt_dec(r.unit_cost) if r.unit_cost is not None else "",
                "Movement Value": (
                    _fmt_dec(r.movement_value) if r.movement_value is not None else ""
                ),
                "Reference Type": r.reference_type or "",
                "Reason": r.reason or "",
                "Notes": r.notes or "",
            }
            for r in rows_raw
        ]
        return (
            _build_xlsx(("STOCK MOVEMENTS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("STOCK MOVEMENTS", headers, rows))
        )

    # export 14: customer list

    async def export_customers(
        self,
        tenant_id: uuid.UUID,
        fmt: str = "csv",
    ) -> bytes:
        order_stats = (
            select(
                Order.customer_id.label("customer_id"),
                func.count(Order.id).label("order_count"),
                func.sum(Order.total_amount).label("total_spent"),
            )
            .where(
                and_(
                    Order.tenant_id == tenant_id,
                    Order.order_status.in_([
                        OrderStatus.COMPLETED.value,
                        OrderStatus.PARTIALLY_REFUNDED.value,
                        OrderStatus.REFUNDED.value,
                    ]),
                )
            )
            .group_by(Order.customer_id)
            .subquery("order_stats")
        )

        stmt = (
            select(
                Customer.customer_code,
                Customer.name,
                Customer.phone,
                Customer.email,
                Customer.gender,
                Customer.current_balance,
                Customer.credit_limit,
                Customer.is_active,
                Customer.created_at,
                func.coalesce(order_stats.c.order_count, literal(0)).label("order_count"),
                func.coalesce(
                    order_stats.c.total_spent, literal(Decimal("0"))
                ).label("total_spent"),
            )
            .outerjoin(order_stats, order_stats.c.customer_id == Customer.id)
            .where(Customer.tenant_id == tenant_id)
            .order_by(Customer.name)
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()
        headers = [
            "Code", "Name", "Phone", "Email", "Gender",
            "Balance", "Credit Limit", "Total Orders", "Total Spent",
            "Status", "Member Since",
        ]
        rows = [
            {
                "Code": r.customer_code or "",
                "Name": r.name,
                "Phone": r.phone or "",
                "Email": r.email or "",
                "Gender": r.gender or "",
                "Balance": _fmt_dec(r.current_balance),
                "Credit Limit": (
                    _fmt_dec(r.credit_limit) if r.credit_limit is not None else ""
                ),
                "Total Orders": str(r.order_count),
                "Total Spent": _fmt_dec(r.total_spent),
                "Status": "Active" if r.is_active else "Inactive",
                "Member Since": _fmt_dt(r.created_at)[:10] if r.created_at else "",
            }
            for r in rows_raw
        ]
        return (
            _build_xlsx(("CUSTOMERS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("CUSTOMERS", headers, rows))
        )

    # export 15: purchase order line items

    async def export_purchase_orders(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        fmt: str = "csv",
    ) -> bytes:
        filters: list = [
            PurchaseOrder.tenant_id == tenant_id,
            PurchaseOrder.deleted_at.is_(None),
        ]
        if start_date:
            filters.append(PurchaseOrder.order_date >= _utc_start(start_date))
        if end_date:
            filters.append(
                PurchaseOrder.order_date < _utc_start(end_date) + timedelta(days=1)
            )

        stmt = (
            select(
                PurchaseOrder.po_number,
                PurchaseOrder.order_date,
                Supplier.name.label("supplier_name"),
                Branch.name.label("branch_name"),
                PurchaseOrder.status,
                Product.name.label("product_name"),
                ProductVariant.name.label("variant_name"),
                func.coalesce(ProductVariant.sku, Product.sku).label("sku"),
                PurchaseOrderItem.ordered_quantity,
                PurchaseOrderItem.received_quantity,
                PurchaseOrderItem.unit_cost,
                PurchaseOrderItem.line_total,
                PurchaseOrder.expected_date,
                PurchaseOrder.notes,
            )
            .join(Supplier, Supplier.id == PurchaseOrder.supplier_id)
            .join(Branch, Branch.id == PurchaseOrder.branch_id)
            .join(PurchaseOrderItem, PurchaseOrderItem.purchase_order_id == PurchaseOrder.id)
            .join(Product, Product.id == PurchaseOrderItem.product_id)
            .outerjoin(ProductVariant, ProductVariant.id == PurchaseOrderItem.variant_id)
            .where(and_(*filters))
            .order_by(
                PurchaseOrder.order_date.desc(),
                PurchaseOrder.po_number,
                Product.name,
            )
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()
        headers = [
            "PO Number", "Order Date", "Supplier", "Branch", "Status",
            "Product", "Variant", "SKU",
            "Ordered Qty", "Received Qty", "Unit Cost", "Line Total",
            "Expected Date", "Notes",
        ]
        rows = [
            {
                "PO Number": r.po_number,
                "Order Date": _fmt_dt(r.order_date)[:10] if r.order_date else "",
                "Supplier": r.supplier_name or "",
                "Branch": r.branch_name or "",
                "Status": r.status,
                "Product": r.product_name,
                "Variant": r.variant_name or "",
                "SKU": r.sku or "",
                "Ordered Qty": _fmt_dec(r.ordered_quantity),
                "Received Qty": _fmt_dec(r.received_quantity),
                "Unit Cost": _fmt_dec(r.unit_cost),
                "Line Total": _fmt_dec(r.line_total),
                "Expected Date": _fmt_dt(r.expected_date)[:10] if r.expected_date else "",
                "Notes": r.notes or "",
            }
            for r in rows_raw
        ]
        return (
            _build_xlsx(("PURCHASE ORDERS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("PURCHASE ORDERS", headers, rows))
        )

    # export 16: goods receipt line items

    async def export_goods_receipts(
        self,
        tenant_id: uuid.UUID,
        start_date: date | None = None,
        end_date: date | None = None,
        fmt: str = "csv",
    ) -> bytes:
        filters: list = [GoodsReceipt.tenant_id == tenant_id]
        if start_date:
            filters.append(GoodsReceipt.receipt_date >= _utc_start(start_date))
        if end_date:
            filters.append(
                GoodsReceipt.receipt_date < _utc_start(end_date) + timedelta(days=1)
            )

        stmt = (
            select(
                GoodsReceipt.receipt_number,
                GoodsReceipt.receipt_date,
                PurchaseOrder.po_number,
                Supplier.name.label("supplier_name"),
                Branch.name.label("branch_name"),
                GoodsReceipt.status,
                Product.name.label("product_name"),
                ProductVariant.name.label("variant_name"),
                func.coalesce(ProductVariant.sku, Product.sku).label("sku"),
                GoodsReceiptItem.received_quantity,
                GoodsReceiptItem.unit_cost,
                GoodsReceiptItem.line_total,
                GoodsReceipt.notes,
            )
            .join(PurchaseOrder, PurchaseOrder.id == GoodsReceipt.purchase_order_id)
            .join(Supplier, Supplier.id == PurchaseOrder.supplier_id)
            .join(Branch, Branch.id == GoodsReceipt.branch_id)
            .join(GoodsReceiptItem, GoodsReceiptItem.goods_receipt_id == GoodsReceipt.id)
            .join(
                PurchaseOrderItem,
                PurchaseOrderItem.id == GoodsReceiptItem.purchase_order_item_id,
            )
            .join(Product, Product.id == PurchaseOrderItem.product_id)
            .outerjoin(ProductVariant, ProductVariant.id == PurchaseOrderItem.variant_id)
            .where(and_(*filters))
            .order_by(
                GoodsReceipt.receipt_date.desc(),
                GoodsReceipt.receipt_number,
                Product.name,
            )
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()
        headers = [
            "Receipt Number", "Receipt Date", "PO Reference", "Supplier", "Branch", "Status",
            "Product", "Variant", "SKU",
            "Received Qty", "Unit Cost", "Line Total",
            "Notes",
        ]
        rows = [
            {
                "Receipt Number": r.receipt_number,
                "Receipt Date": _fmt_dt(r.receipt_date)[:10] if r.receipt_date else "",
                "PO Reference": r.po_number or "",
                "Supplier": r.supplier_name or "",
                "Branch": r.branch_name or "",
                "Status": r.status,
                "Product": r.product_name,
                "Variant": r.variant_name or "",
                "SKU": r.sku or "",
                "Received Qty": _fmt_dec(r.received_quantity),
                "Unit Cost": _fmt_dec(r.unit_cost),
                "Line Total": _fmt_dec(r.line_total),
                "Notes": r.notes or "",
            }
            for r in rows_raw
        ]
        return (
            _build_xlsx(("GOODS RECEIPTS", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("GOODS RECEIPTS", headers, rows))
        )

    # export 17: supplier payables

    async def export_supplier_payables(
        self,
        tenant_id: uuid.UUID,
        fmt: str = "csv",
    ) -> bytes:
        stmt = (
            select(
                Supplier.name.label("supplier_name"),
                PurchaseOrder.po_number,
                PurchaseOrder.order_date,
                Branch.name.label("branch_name"),
                SupplierPayable.total_amount,
                SupplierPayable.paid_amount,
                SupplierPayable.remaining_amount,
                SupplierPayable.status,
            )
            .join(Supplier, Supplier.id == SupplierPayable.supplier_id)
            .join(PurchaseOrder, PurchaseOrder.id == SupplierPayable.purchase_order_id)
            .join(Branch, Branch.id == PurchaseOrder.branch_id)
            .where(SupplierPayable.tenant_id == tenant_id)
            .order_by(SupplierPayable.status, Supplier.name, PurchaseOrder.order_date.desc())
        )

        rows_raw = (await self.session.execute(stmt)).mappings().all()
        headers = [
            "Supplier", "PO Number", "PO Date", "Branch",
            "Invoice Amount", "Paid Amount", "Remaining Amount", "Status",
        ]
        rows = [
            {
                "Supplier": r.supplier_name or "",
                "PO Number": r.po_number,
                "PO Date": _fmt_dt(r.order_date)[:10] if r.order_date else "",
                "Branch": r.branch_name or "",
                "Invoice Amount": _fmt_dec(r.total_amount),
                "Paid Amount": _fmt_dec(r.paid_amount),
                "Remaining Amount": _fmt_dec(r.remaining_amount),
                "Status": r.status,
            }
            for r in rows_raw
        ]
        return (
            _build_xlsx(("SUPPLIER PAYABLES", headers, rows))
            if fmt == "xlsx"
            else _build_csv(("SUPPLIER PAYABLES", headers, rows))
        )
